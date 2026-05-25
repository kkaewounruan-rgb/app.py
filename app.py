import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="ระบบสรุปรายงานประเมินการอบรม", layout="wide")

st.title("📊 ระบบจัดทำรายงานสรุปการประเมินหลังการอบรม")
st.markdown("นำเข้าข้อมูลจากไฟล์ CSV/Excel หรือ ลิงก์ Google Sheet เพื่อสร้างรายงานสรุป")

# 1. เลือกแหล่งข้อมูล
source_type = st.radio("เลือกแหล่งข้อมูลนำเข้า", ("อัปโหลดไฟล์ข้อมูล (.csv, .xlsx)", "ลิงก์ Google Sheet"))

df = None

if source_type == "อัปโหลดไฟล์ข้อมูล (.csv, .xlsx)":
    uploaded_file = st.file_uploader("อัปโหลดไฟล์การตอบกลับ", type=["csv", "xlsx"])
    if uploaded_file is not None:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)

elif source_type == "ลิงก์ Google Sheet":
    gsheet_url = st.text_input("วางลิงก์ Google Sheet (ต้องตั้งค่าแชร์เป็น Anyone with the link)")
    if gsheet_url:
        try:
            # แปลงลิงก์เพื่อให้โหลดเป็น CSV ได้
            csv_url = gsheet_url.replace('/edit?usp=sharing', '/export?format=csv').replace('/edit#gid=', '/export?format=csv&gid=')
            df = pd.read_csv(csv_url)
            st.success("นำเข้าข้อมูลสำเร็จ!")
        except Exception as e:
            st.error(f"ไม่สามารถดึงข้อมูลได้ โปรดตรวจสอบลิงก์และสิทธิ์การเข้าถึง (Error: {e})")

# 2. ประมวลผลและสร้างไฟล์
if df is not None:
    st.write(f"จำนวนผู้ตอบแบบประเมินทั้งหมด: **{len(df)}** คน")
    st.dataframe(df.head(5)) # แสดงตัวอย่างข้อมูล
    
    if st.button("📥 สร้างไฟล์รายงานสรุป (Excel/PDF Ready)"):
        # สร้าง Workbook แบบฟอร์ม
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Report"
        
        # ตั้งค่าฟอนต์และสไตล์
        th_font = Font(name='Tahoma', size=11)
        header_font = Font(name='Tahoma', size=11, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.append(["สรุปผลประเมินการอบรม"])
        ws.append(["จำนวนผู้ประเมินทั้งสิ้น:", len(df), "คน"])
        ws.append([])
        
        headers = ["หัวข้อประเมิน", "มากที่สุด (5)", "มาก (4)", "ปานกลาง (3)", "น้อย (2)", "น้อยที่สุด (1)", "เฉลี่ย"]
        ws.append(headers)
        
        # ค้นหาคอลัมน์ที่เป็นคำถามประเมิน (มีตัวเลข 1-5)
        question_cols = [col for col in df.columns if df[col].dtype in ['int64', 'float64'] and col not in ['ประทับเวลา', 'รหัสพนักงาน']]
        
        for q in question_cols:
            counts = df[q].value_counts()
            avg = df[q].mean() if not df[q].isna().all() else 0
            
            row_data = [
                q[:80] + '...' if len(q) > 80 else q, # ตัดชื่อให้สั้นลง
                counts.get(5, 0), counts.get(4, 0), counts.get(3, 0),
                counts.get(2, 0), counts.get(1, 0), round(avg, 2)
            ]
            ws.append(row_data)
            
        # จัด Format ให้สวยงาม
        for row in ws.iter_rows(min_row=4, max_col=7, max_row=ws.max_row):
            for cell in row:
                cell.font = th_font
                cell.border = thin_border
                if cell.column > 1:
                    cell.alignment = center_align
        ws.column_dimensions['A'].width = 60
        
        # แปลงไฟล์เพื่อเตรียม Download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(
            label="⬇️ ดาวน์โหลดรายงานสรุป (นำไป Save as เป็น PDF หรือ Print ได้ทันที)",
            data=output,
            file_name="Summary_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )